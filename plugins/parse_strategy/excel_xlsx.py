# pylint: disable=W0613
""" Strategy class for workbook/xlsx """
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, DirectoryPath

from app.models.resourceconfig import ResourceConfig
from app.strategy.factory import StrategyFactory


class XLSXParseDataModel(BaseModel):
    """Data model for retrieving a rectangular section of an Excel sheet.

    Fields:
        worksheet: Name of worksheet to load.
        row_from: Excel row number of first row.  Defaults to first
          assigned row.
        col_from: Excel column number or label of first column.
          Defaults to first assigned column.
        row_to: Excel row number of last row.  Defaults to last
          assigned row.
        col_to: Excel column number or label of last column.  Defaults
          to last assigned column.
        headerRow: Row number with the headers. Defaults to 1 if
          header is given, otherwise None.
        header: Optional list of column names, specifying the columns
          to return.  These names they should match cells in `headerRow`.
        newHeader: Optional list of new column names replacing `header`
          in the output.
        downloadDir: Directory where we expect to find the (possible
          downloaded) excel file.  Defaults to the current directory,
          unless the 'OTEAPI_downloadDir' environment variable is set.
    """

    worksheet: str
    row_from: int = None
    col_from: Union[int, str] = None
    row_to: int = None
    col_to: Union[int, str] = None
    headerRow: int = None
    header: List[str] = None
    newHeader: List[str] = None
    downloadDir: DirectoryPath = (  # move to ResourceConfig??
        os.environ["OTEAPI_downloadDir"] if "OTEAPI_downloadDir" in os.environ else "."
    )


def set_model_defaults(model: XLSXParseDataModel, worksheet: Worksheet):
    """Update datamodel `model` with default values obtained from `worksheet`."""
    if model.row_from is None:
        if model.header:
            # assume that data starts on the first row after the header
            model.row_from = model.headerRow + 1
        else:
            model.row_from = worksheet.min_row

    if model.row_to is None:
        model.row_to = worksheet.max_row

    if model.col_from is None:
        model.col_from = worksheet.min_column
    elif isinstance(model.col_from, str):
        model.col_from = column_index_from_string(model.col_from)

    if model.col_to is None:
        model.col_to = worksheet.max_column
    elif isinstance(model.col_to, str):
        model.col_to = column_index_from_string(model.col_to)

    if model.header and not model.headerRow:
        model.headerRow = 1


def get_column_indices(model: XLSXParseDataModel, worksheet: Worksheet) -> List[int]:
    """Helper function returning a list of column indices."""
    if model.header:
        header_dict = {
            worksheet.cell(model.headerRow, col).value: col
            for col in range(model.col_from, model.col_to + 1)
        }
        indices = [header_dict[h] for h in model.header]
    else:
        indices = range(model.col_from, model.col_to + 1)
    return indices


@dataclass
@StrategyFactory.register(
    ("mediaType", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
)
class XLSXParseStrategy:

    resource_config: ResourceConfig

    def __post_init__(self):
        if self.resource_config.downloadUrl.scheme == "file":
            # Workaround for strange behaviour (bug?) in "file" scheme for AnyUrl
            self.path = Path(self.resource_config.downloadUrl.host)
        else:
            self.path = Path(self.resource_config.downloadUrl.path)

        if self.resource_config.configuration:
            self.config = self.resource_config.configuration
        else:
            self.config = {}

    def parse(self, session: Optional[Dict[str, Any]] = None) -> Dict:
        """Parses selected region of an excel file.  The fields of the
        returned dict are as follows:
        - header: list of column headers (read from `headerRow`).  Is None
            if neither `headerRow` nor `header` configuration is given.
        - data: cell values (as a list of lists, row-wise)
        """
        model = XLSXParseDataModel(**self.config)
        filename = Path(model.downloadDir) / self.path.name
        workbook = load_workbook(filename=filename, read_only=True, data_only=True)
        worksheet = workbook[model.worksheet]
        set_model_defaults(model, worksheet)
        columns = get_column_indices(model, worksheet)

        data = []
        for row in worksheet.iter_rows(
            min_row=model.row_from,
            max_row=model.row_to,
            min_col=min(columns),
            max_col=max(columns),
        ):
            data.append([row[c - 1].value for c in columns])

        if model.headerRow:
            row = worksheet.iter_rows(
                min_row=model.headerRow,
                max_row=model.headerRow,
                min_col=min(columns),
                max_col=max(columns),
            ).__next__()
            header = [row[c - 1].value for c in columns]
        else:
            header = None

        if model.newHeader:
            nhead = len(header) if header else len(data[0]) if data else 0
            if len(model.newHeader) != nhead:
                raise TypeError(
                    f"length of `newHeader` (={len(model.newHeader)}) "
                    f"doesn't match number of columns (={len(header)})"
                )
            if header:
                for i, val in enumerate(model.newHeader):
                    if val is not None:
                        header[i] = val
            elif data:
                header = model.newHeader

        return {"header": header, "data": data}

    def parse_recarray(self, session: Optional[Dict[str, Any]] = None) -> Dict:
        """Convenient method that works like `parse()`, but returns the data
        as a numpy recarray.
        """
        import numpy as np

        d = self.parse(session)
        data = d["data"]
        nhead = len(d["header"]) if d["header"] else 0
        nrows = len(data)
        ncols = len(data[0]) if nrows else nhead

        # Set blank cells in numerical columns to NaN such that
        # np.rec.fromrecords() will convert them to the correct
        # numerical dtype
        for col in range(ncols):
            if all(
                isinstance(data[row][col], (bool, int, float, complex, None.__class__))
                for row in range(nrows)
            ):
                for row in range(nrows):
                    if data[row][col] is None:
                        data[row][col] = np.nan

        d["data"] = np.rec.fromrecords(data, names=d["header"])
        return d
