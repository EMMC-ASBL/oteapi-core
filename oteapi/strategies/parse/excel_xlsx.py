"""Strategy class for workbook/xlsx."""

from __future__ import annotations

import sys
from typing import TYPE_CHECKING, Optional, Union

if sys.version_info >= (3, 10):
    from typing import Literal
else:
    from typing_extensions import Literal

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import Field
from pydantic.dataclasses import dataclass

from oteapi.datacache import DataCache
from oteapi.models import AttrDict, DataCacheConfig, ParserConfig, ResourceConfig
from oteapi.plugins import create_strategy

if TYPE_CHECKING:  # pragma: no cover
    from collections.abc import Iterable

    from openpyxl.worksheet.worksheet import Worksheet

from oteapi.models.resourceconfig import HostlessAnyUrl


class XLSXParseContent(AttrDict):
    """Class for returning values from XLSXParse."""

    data: dict[str, list] = Field(
        ...,
        description="A dict with column-name/column-value pairs. The values are lists.",
    )


class XLSXParseConfig(AttrDict):
    """Data model for retrieving a rectangular section of an Excel sheet."""

    # Resource config
    downloadUrl: Optional[HostlessAnyUrl] = Field(
        None, description=ResourceConfig.model_fields["downloadUrl"].description
    )
    mediaType: Literal[
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ] = Field(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        description=ResourceConfig.model_fields["mediaType"].description,
    )

    # XLSX parse strategy-specific config
    worksheet: str = Field(..., description="Name of worksheet to load.")
    row_from: Optional[int] = Field(
        None,
        description="Excel row number of first row. Defaults to first assigned row.",
    )
    col_from: Optional[Union[int, str]] = Field(
        None,
        description=(
            "Excel column number or label of first column. Defaults to first assigned "
            "column."
        ),
    )
    row_to: Optional[int] = Field(
        None, description="Excel row number of last row. Defaults to last assigned row."
    )
    col_to: Optional[Union[int, str]] = Field(
        None,
        description=(
            "Excel column number or label of last column. Defaults to last assigned "
            "column."
        ),
    )
    header_row: Optional[int] = Field(
        None,
        description=(
            "Row number with the headers. Defaults to `1` if header is given, "
            "otherwise `None`."
        ),
    )
    header: Optional[list[str]] = Field(
        None,
        description=(
            "Optional list of column names, specifying the columns to return. "
            "These names they should match cells in `header_row`."
        ),
    )
    new_header: Optional[list[str]] = Field(
        None,
        description=(
            "Optional list of new column names replacing `header` in the output."
        ),
    )
    download_config: AttrDict = Field(
        AttrDict(),
        description="Configurations provided to a download strategy.",
    )
    datacache_config: Optional[DataCacheConfig] = Field(
        None,
        description=(
            "Configurations for the data cache for retrieving the downloaded file "
            "content."
        ),
    )


class XLSXParseParserConfig(ParserConfig):
    """XLSX parse strategy resource config."""

    parserType: Literal["parser/excel_xlsx"] = Field(
        "parser/excel_xlsx",
        description=ParserConfig.model_fields["parserType"].description,
    )
    configuration: XLSXParseConfig = Field(
        ..., description="SQLite parse strategy-specific configuration."
    )


def set_model_defaults(model: XLSXParseConfig, worksheet: Worksheet) -> None:
    """Update data model `model` with default values obtained from `worksheet`.

    Parameters:
        model: The parsed data model.
        worksheet: Excel worksheet, from which the default values will be obtained.

    """
    if model.row_from is None:
        if model.header:
            # assume that data starts on the first row after the header
            model.row_from = model.header_row + 1 if model.header_row else 1
        else:
            model.row_from = worksheet.min_row

    if model.row_to is None:
        model.row_to = worksheet.max_row

    if model.col_from is None:
        model.col_from = worksheet.min_column
    elif isinstance(model.col_from, str):
        model.col_from = column_index_from_string(model.col_from)

    if model.col_to is None:
        model.col_to = worksheet.max_column
    elif isinstance(model.col_to, str):
        model.col_to = column_index_from_string(model.col_to)

    if model.header and not model.header_row:
        model.header_row = 1


def get_column_indices(model: XLSXParseConfig, worksheet: Worksheet) -> Iterable[int]:
    """Helper function returning a list of column indices.

    Parameters:
        model: The parsed data model.
        worksheet: Excel worksheet, from which the header values will be retrieved.

    Returns:
        A list of column indices.

    """
    if not isinstance(model.col_from, int) or not isinstance(model.col_to, int):
        raise TypeError("Expected `model.col_from` and `model.col_to` to be integers.")

    if model.header:
        header_dict = {
            worksheet.cell(model.header_row, col).value: col
            for col in range(model.col_from, model.col_to + 1)
        }
        return [header_dict[h] for h in model.header]
    return range(model.col_from, model.col_to + 1)


@dataclass
class XLSXParseStrategy:
    """Parse strategy for Excel XLSX files.

    **Registers strategies**:

    - `("parserType", "excel_xlsx")`

    """

    parse_config: XLSXParseParserConfig

    def initialize(self) -> AttrDict:
        """Initialize."""
        return AttrDict()

    def get(self) -> XLSXParseContent:
        """Parses selected region of an excel file.

        Returns:
            A dict with column-name/column-value pairs. The values are lists.

        """

        config = self.parse_config.configuration

        # Download the file
        download_config = config.model_dump()
        download_config["configuration"] = config.download_config.model_dump()
        output = create_strategy("download", download_config).get()

        if config.datacache_config and config.datacache_config.accessKey:
            cache_key = config.datacache_config.accessKey
        elif "key" in output:
            cache_key = output["key"]
        else:
            raise RuntimeError("No data cache key provided to the downloaded content")

        cache = DataCache(config.datacache_config)

        with cache.getfile(key=cache_key, suffix=".xlsx") as filename:
            # Note that we have to set read_only=False to ensure that
            # load_workbook() properly closes the xlsx file after reading.
            # Otherwise Windows will fail when the temporary file is removed
            # when leaving the with statement.
            workbook = load_workbook(filename=filename, read_only=False, data_only=True)

        worksheet = workbook[config.worksheet]
        set_model_defaults(config, worksheet)
        columns = get_column_indices(config, worksheet)

        data = []
        for row in worksheet.iter_rows(
            min_row=config.row_from,
            max_row=config.row_to,
            min_col=min(columns),
            max_col=max(columns),
        ):
            data.append([row[c - 1].value for c in columns])

        if config.header_row:
            row = worksheet.iter_rows(
                min_row=config.header_row,
                max_row=config.header_row,
                min_col=min(columns),
                max_col=max(columns),
            ).__next__()
            header = [row[c - 1].value for c in columns]
        else:
            header = None

        if config.new_header:
            nhead = len(header) if header else len(data[0]) if data else 0
            if len(config.new_header) != nhead:
                raise TypeError(
                    "length of `new_header` "
                    f"(={len(config.new_header)}) "
                    f"doesn't match number of columns (={len(header) if header else 0})"
                )
            if header:
                for i, val in enumerate(config.new_header):
                    if val is not None:
                        header[i] = val
            elif data:
                header = config.new_header

        if header is None:
            header = [get_column_letter(col + 1) for col in range(len(data))]

        transposed = [list(datum) for datum in zip(*data)]
        return XLSXParseContent(data=dict(zip(header, transposed)))
