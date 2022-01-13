# pylint: disable=W0613
""" Strategy class for workbook/xlsx """
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from oteapi.datacache.datacache import DataCache
from oteapi.models.resourceconfig import ResourceConfig
from oteapi.strategy-interfaces.factory import StrategyFactory
from oteapi.strategy-interfaces.idownloadstrategy import create_download_strategy
from pydantic import BaseModel, Extra


class XLSXParseDataModel(BaseModel):
    """Data model for retrieving a rectangular section of an Excel sheet.

    Fields:
        worksheet: Name of worksheet to load.
        row_from: Excel row number of first row.  Defaults to first
          assigned row.
        col_from: Excel column number or label of first column.
          Defaults to first assigned column.
        row_to: Excel row number of last row.  Defaults to last
          assigned row.
        col_to: Excel column number or label of last column.  Defaults
          to last assigned column.
        header_row: Row number with the headers. Defaults to 1 if
          header is given, otherwise None.
        header: Optional list of column names, specifying the columns
          to return.  These names they should match cells in `header_row`.
        new_header: Optional list of new column names replacing `header`
          in the output.
    """

    worksheet: str
    row_from: int = None
    col_from: Union[int, str] = None
    row_to: int = None
    col_to: Union[int, str] = None
    header_row: int = None
    header: List[str] = None
    new_header: List[str] = None


def set_model_defaults(model: XLSXParseDataModel, worksheet: Worksheet):
    """Update datamodel `model` with default values obtained from `worksheet`."""
    if model.row_from is None:
        if model.header:
            # assume that data starts on the first row after the header
            model.row_from = model.header_row + 1
        else:
            model.row_from = worksheet.min_row

    if model.row_to is None:
        model.row_to = worksheet.max_row

    if model.col_from is None:
        model.col_from = worksheet.min_column
    elif isinstance(model.col_from, str):
        model.col_from = column_index_from_string(model.col_from)

    if model.col_to is None:
        model.col_to = worksheet.max_column
    elif isinstance(model.col_to, str):
        model.col_to = column_index_from_string(model.col_to)

    if model.header and not model.header_row:
        model.header_row = 1


def get_column_indices(model: XLSXParseDataModel, worksheet: Worksheet) -> List[int]:
    """Helper function returning a list of column indices."""
    if model.header:
        header_dict = {
            worksheet.cell(model.header_row, col).value: col
            for col in range(model.col_from, model.col_to + 1)
        }
        indices = [header_dict[h] for h in model.header]
    else:
        indices = range(model.col_from, model.col_to + 1)
    return indices


@dataclass
@StrategyFactory.register(
    ("mediaType", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
)
class XLSXParseStrategy:

    resource_config: ResourceConfig

    def initialize(
        self, session: Optional[Dict[str, Any]] = None  # pylint: disable=W0613
    ) -> Dict:
        """Initialize"""
        return {}

    def parse(self, session: Optional[Dict[str, Any]] = None) -> Dict:
        """Parses selected region of an excel file.

        Returns a dict with column-name/column-value pairs.  The values are
        lists.
        """
        model = XLSXParseDataModel(
            **self.resource_config.configuration, extra=Extra.ignore
        )

        downloader = create_download_strategy(self.resource_config)
        output = downloader.get()
        cache = DataCache(self.resource_config.configuration)
        with cache.getfile(key=output["key"], suffix=".xlsx") as filename:
            workbook = load_workbook(filename=filename, read_only=True, data_only=True)
        worksheet = workbook[model.worksheet]
        set_model_defaults(model, worksheet)
        columns = get_column_indices(model, worksheet)

        data = []
        for row in worksheet.iter_rows(
            min_row=model.row_from,
            max_row=model.row_to,
            min_col=min(columns),
            max_col=max(columns),
        ):
            data.append([row[c - 1].value for c in columns])

        if model.header_row:
            row = worksheet.iter_rows(
                min_row=model.header_row,
                max_row=model.header_row,
                min_col=min(columns),
                max_col=max(columns),
            ).__next__()
            header = [row[c - 1].value for c in columns]
        else:
            header = None

        if model.new_header:
            nhead = len(header) if header else len(data[0]) if data else 0
            if len(model.new_header) != nhead:
                raise TypeError(
                    f"length of `new_header` (={len(model.new_header)}) "
                    f"doesn't match number of columns (={len(header)})"
                )
            if header:
                for i, val in enumerate(model.new_header):
                    if val is not None:
                        header[i] = val
            elif data:
                header = model.new_header

        if header is None:
            header = [get_column_letter(col + 1) for col in range(len(data))]

        transposed = list(map(list, zip(*data)))
        return {k: v for k, v in zip(header, transposed)}
